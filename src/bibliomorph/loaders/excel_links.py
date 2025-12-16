from enum import Enum
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd
from typing import Callable


from .loader import BaseLoader


# class NameSource(Enum):
#     COLUMN = 0
#     SHEET = 1
#     KEY = 2


class ExcelLinksLoader(BaseLoader):

    skip_sheets: list[str] = []
    source: str  #  | NameSource
    target: str  #  | NameSource
    source_formatter: Callable[[list[str]], list[str]]
    target_formatter: Callable[[list[str]], list[str]]

    def load(self, path: Path):
        data = load_workbook(path)
        sheets = {}
        for sheet_name in data.sheetnames:
            if sheet_name in self.skip_sheets:
                continue
            sheets[sheet_name] = pd.read_excel(path, sheet_name)

        links = set()

        for sheet_name, sheet in sheets.items():
            sources = self.source_formatter(sheet[self.source])
            targets = self.target_formatter(sheet[self.target])

            links = links.union(set(zip(sources, targets)))

        return [], [{"source": source, "target": target} for source, target in links]
